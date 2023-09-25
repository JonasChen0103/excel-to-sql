import openpyxl

# Set your Excel file name here
source_file = 'your-file-name.xls'
workbook = openpyxl.load_workbook(source_file)

# Set your worksheet here.
# For the first worksheet, type '1'
worksheet = 1
sheet = workbook.worksheets[worksheet - 1]

# Open a text file in write mode with UTF-8 encoding
# Set your output file name here
output_file = 'your-output-file-name.txt'
with open(output_file, 'w', encoding='utf-8') as file:
    # Generate SQL and write it to the file
    # Set your table name here
    table_name = 'your-table-name'
    for i in range(2, sheet.max_row + 1):
        file.write(f"-- ROW_NO: {i}\n")
        file.write(f"DELETE FROM {table_name}\n")
        file.write('WHERE 1 = 1\n')
        for j in range(2, sheet.max_column + 1):
            column_name = sheet.cell(row=1, column=j).value
            # Convert None to an empty string
            cell_data = sheet.cell(row=i, column=j).value
            cell_data = cell_data if cell_data is not None else ''
            if cell_data == 'NULL':
                file.write(f"AND {column_name} IS {cell_data}\n")
            else:
                file.write(f"AND {column_name} = '{cell_data}'\n")
        file.write('\n')

print("SQL statements have been successfully written to the output.txt file.")
